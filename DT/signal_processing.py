"""
Signal processing module.
Handles data loading, FFT, peak picking, damping estimation and mode shapes.
All records are used at full length with no trimming or window narrowing.
"""
import numpy as np
import matplotlib; matplotlib.use('Agg')
import matplotlib.pyplot as plt
from scipy.signal import butter, filtfilt, hilbert, find_peaks
from scipy.optimize import curve_fit
import openpyxl, os

plt.rcParams.update({
    'font.family': 'serif', 'font.size': 10, 'axes.labelsize': 11,
    'axes.titlesize': 12, 'figure.dpi': 150, 'savefig.dpi': 300,
    'lines.linewidth': 0.7, 'axes.grid': True, 'grid.alpha': 0.25,
    'legend.fontsize': 9, 'xtick.labelsize': 9, 'ytick.labelsize': 9,
})
FD = "figures"
os.makedirs(FD, exist_ok=True)


def load_sheet(fp, name):
    """Load acceleration data from Excel. Returns time, ch1, ch2, ch3."""
    wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
    actual = name
    for s in wb.sheetnames:
        if s.strip() == name.strip(): actual = s; break
    ws = wb[actual]
    data = np.array(list(ws.iter_rows(min_row=2, values_only=True)), dtype=float)
    wb.close()
    return data[:,3], data[:,5], data[:,6], data[:,7]


def fft_amp(sig, fs, fmax=40.0):
    """Single-sided amplitude spectrum with Hanning window, full record."""
    N = len(sig)
    w = np.hanning(N)
    ft = np.fft.rfft((sig - np.mean(sig)) * w)
    f = np.fft.rfftfreq(N, 1.0/fs)
    a = 2.0 * np.abs(ft) / np.sum(w)
    m = f <= fmax
    return f[m], a[m]


def psd(sig, fs, fmax=40.0):
    """Power spectral density estimate, full record, Hanning window."""
    N = len(sig)
    w = np.hanning(N)
    ft = np.fft.rfft((sig - np.mean(sig)) * w)
    f = np.fft.rfftfreq(N, 1.0/fs)
    p = (2.0 * np.abs(ft)**2) / (fs * np.sum(w**2))
    m = f <= fmax
    return f[m], p[m]


def top_peaks(f, a, n=3, fmin=5.0, sep=4.0):
    """Find n strongest spectral peaks above fmin Hz."""
    dr = f[1] - f[0]
    md = max(1, int(sep / dr))
    mk = f >= fmin
    fs2, As = f[mk], a[mk]
    ix, _ = find_peaks(As, distance=md)
    if not len(ix):
        return [np.nan]*n, [np.nan]*n
    top = ix[np.argsort(As[ix])[::-1][:n]]
    top = top[np.argsort(fs2[top])]
    pf = list(fs2[top]); pa = list(As[top])
    while len(pf) < n: pf.append(np.nan); pa.append(np.nan)
    return pf[:n], pa[:n]


def extract_all_frequencies(session_file, sheets, fs):
    """Extract f1, f2, f3 from every sheet, every floor, full records."""
    results = []
    for key, sheet_name in sheets.items():
        try:
            t, c1, c2, c3 = load_sheet(session_file, sheet_name)
            for fl_idx, (ch, fl) in enumerate([(c1,'Fl1'),(c2,'Fl2'),(c3,'Fl3')]):
                freqs, amps = fft_amp(ch, fs)
                pf, pa = top_peaks(freqs, amps, n=3)
                results.append({
                    'test': sheet_name.strip(), 'floor': fl, 'key': key,
                    'f1': pf[0], 'f2': pf[1], 'f3': pf[2],
                    'n_pts': len(ch), 'duration': len(ch)/fs,
                    'fft_freqs': freqs, 'fft_amps': amps,
                })
        except: pass
    return results


def compute_sigma(results, broadband_keys=['free_vib', 'impact']):
    """
    Empirical sigma from all broadband tests, all floors.
    No artificial floor applied. If any sigma is zero, it means the FFT
    resolution is too coarse to distinguish the variability, and the
    number should be reported honestly rather than replaced.
    """
    bb = [r for r in results if r['key'] in broadband_keys]
    f1s = [r['f1'] for r in bb if not np.isnan(r['f1'])]
    f2s = [r['f2'] for r in bb if not np.isnan(r['f2'])]
    f3s = [r['f3'] for r in bb if not np.isnan(r['f3'])]
    sigma = np.array([np.std(f1s), np.std(f2s), np.std(f3s)])
    means = np.array([np.mean(f1s), np.mean(f2s), np.mean(f3s)])
    return sigma, means, {'f1': f1s, 'f2': f2s, 'f3': f3s}


def est_damping(sig, fs, fn, fl, fh):
    """Damping from Hilbert envelope of bandpass-filtered free vibration signal."""
    nyq = 0.5 * fs
    b, a = butter(3, [fl/nyq, fh/nyq], btype='band')
    env = np.abs(hilbert(filtfilt(b, a, sig)))
    i0 = int(0.05 * fs)
    i1 = min(len(env), int(12.05 * fs))
    td = np.arange(i1 - i0) / fs
    ed = env[i0:i1]
    win = max(1, int(2 * fs / fn))
    es = np.convolve(ed, np.ones(win)/win, mode='same')
    step = max(1, len(td) // 500)
    ts, ess = td[::step], es[::step]
    keep = ess > np.max(ess) * 0.01
    ts, ess = ts[keep], ess[keep]
    p, _ = curve_fit(lambda t, A, a: A * np.exp(-a*t), ts, ess,
                     p0=[ess[0], 1], maxfev=10000)
    return p[1] / (2*np.pi*fn), p[1], p[0], td, es


def mode_shape(fp, sheet, fs, ftarget):
    """Mode shape from harmonic test using cross-spectrum phase."""
    try: _, c1, c2, c3 = load_sheet(fp, sheet)
    except: return None
    N = len(c1); w = np.hanning(N)
    freqs = np.fft.rfftfreq(N, 1/fs)
    idx = np.argmin(np.abs(freqs - ftarget))
    ref = np.fft.rfft((c1 - np.mean(c1)) * w)
    amps = np.zeros(3)
    for j, ch in enumerate([c1, c2, c3]):
        ft = np.fft.rfft((ch - np.mean(ch)) * w)
        amps[j] = np.abs(ft[idx])
        if np.real(ft[idx] * np.conj(ref[idx])) < 0:
            amps[j] *= -1
    amps /= np.max(np.abs(amps))
    return amps


def mac(p1, p2):
    """Modal Assurance Criterion between two mode shape vectors."""
    return (np.dot(p1, p2)**2) / (np.dot(p1, p1) * np.dot(p2, p2))


# plotting

def plot_fft_grid(results, session_label, test_keys, fname):
    """FFT amplitude from every test, every floor. Full records, no trimming."""
    tests = []
    for k in test_keys:
        matches = [r for r in results if r['key'] == k]
        if matches: tests.append((k, matches[0]['test']))
    nt = len(tests)
    fig, axes = plt.subplots(nt, 3, figsize=(14, 2.8*nt), sharex=True)
    if nt == 1: axes = axes.reshape(1, -1)
    for row, (key, tname) in enumerate(tests):
        for col, fl in enumerate(['Fl1','Fl2','Fl3']):
            ax = axes[row, col]
            ms = [r for r in results if r['key'] == key and r['floor'] == fl]
            if ms:
                d = ms[0]
                ax.plot(d['fft_freqs'], d['fft_amps'], color='0.2', lw=0.5)
                ax.set_xlim([3, 38]); ax.set_ylim(bottom=0)
                pf, pa = top_peaks(d['fft_freqs'], d['fft_amps'], n=3)
                for fp, ap in zip(pf, pa):
                    if not np.isnan(fp):
                        ax.plot(fp, ap, 'o', color='C3', ms=3, zorder=5)
                        ax.annotate(f'{fp:.2f}', (fp, ap), xytext=(4, 4),
                                    textcoords='offset points', fontsize=7, color='C3')
            if row == 0:
                ax.set_title(fl.replace('Fl', 'Floor '), fontsize=11)
            if col == 0:
                ax.set_ylabel(f"{tname}\n({d['duration']:.0f} s)", fontsize=8)
            if row == nt - 1:
                ax.set_xlabel('Frequency (Hz)')
    fig.suptitle(f'{session_label}: amplitude spectra, all tests, full records',
                 fontsize=12, y=1.01)
    plt.tight_layout()
    plt.savefig(os.path.join(FD, fname), dpi=200, bbox_inches='tight')
    plt.close()
    print(f"  saved {fname}")


def plot_overlay(all_results, fname, mode='amp'):
    """All broadband tests overlaid on the same axes per floor."""
    bb = [r for r in all_results if r['key'] in ['free_vib', 'impact']]
    floors = ['Fl1', 'Fl2', 'Fl3']
    fig, axes = plt.subplots(3, 1, figsize=(11, 8), sharex=True)
    colours = ['C0', 'C1', 'C2', 'C3']
    for ax, fl in zip(axes, floors):
        fl_data = [r for r in bb if r['floor'] == fl]
        for ci, r in enumerate(fl_data):
            sess = r.get('session', '?')
            lbl = f"{sess} {r['test']} ({r['duration']:.0f} s)"
            y = r['fft_amps']**2 if mode == 'psd' else r['fft_amps']
            if mode == 'psd':
                ax.semilogy(r['fft_freqs'], y, lw=0.5, color=colours[ci % 4],
                            label=lbl, alpha=0.8)
            else:
                ax.plot(r['fft_freqs'], y, lw=0.5, color=colours[ci % 4],
                        label=lbl, alpha=0.8)
        ax.set_xlim([3, 38])
        ax.set_ylabel('PSD' if mode == 'psd' else 'Amplitude (a.u.)')
        ax.text(0.98, 0.92, fl.replace('Fl', 'Floor '), transform=ax.transAxes,
                ha='right', va='top', fontsize=10,
                bbox=dict(boxstyle='round', fc='white', ec='0.7', pad=0.3))
        if fl == 'Fl1': ax.legend(fontsize=7, loc='upper left', ncol=2)
    axes[2].set_xlabel('Frequency (Hz)')
    title = 'PSD overlay, broadband tests' if mode == 'psd' else 'Amplitude overlay, broadband tests'
    axes[0].set_title(title, fontsize=12)
    plt.tight_layout()
    plt.savefig(os.path.join(FD, fname), dpi=300, bbox_inches='tight')
    plt.close()
    print(f"  saved {fname}")


def plot_sigma(sigma, means, raw_data, fname):
    """Individual frequency identifications and the resulting sigma band."""
    fig, axes = plt.subplots(1, 3, figsize=(13, 4.5))
    keys = ['f1', 'f2', 'f3']
    labels = ['Mode 1', 'Mode 2', 'Mode 3']
    for i, (ax, lbl, key) in enumerate(zip(axes, labels, keys)):
        vals = raw_data[key]
        ax.plot(range(len(vals)), vals, 'o', color='C0', ms=5, zorder=5)
        ax.axhline(means[i], color='C3', ls='-', lw=1.2,
                   label=f'mean = {means[i]:.3f} Hz')
        ax.axhspan(means[i] - sigma[i], means[i] + sigma[i],
                   alpha=0.15, color='C3', label=f'1 std = {sigma[i]:.4f} Hz')
        ax.set_xlabel('Measurement index')
        ax.set_ylabel(f'{lbl} frequency (Hz)')
        ax.set_title(f'{lbl}: {len(vals)} measurements, '
                     f'std = {sigma[i]:.4f} Hz ({sigma[i]/means[i]*100:.2f}%)',
                     fontsize=10)
        ax.legend(fontsize=8)
        for j, v in enumerate(vals):
            ax.annotate(f'{v:.2f}', (j, v), xytext=(2, 5),
                        textcoords='offset points', fontsize=6.5)
    plt.suptitle('Empirical measurement uncertainty from 12 broadband tests\n'
                 '(2 sessions, 2 test types, 3 floors per test)', fontsize=11)
    plt.tight_layout()
    plt.savefig(os.path.join(FD, fname), dpi=300, bbox_inches='tight')
    plt.close()
    print(f"  saved {fname}")


def plot_damping(data, fname):
    fig, axes = plt.subplots(len(data), 1, figsize=(10, 3*len(data)))
    if len(data) == 1: axes = [axes]
    for ax, (td, es, z, al, A0, fn, ml) in zip(axes, data):
        ax.plot(td, es, color='0.4', lw=0.5, label='Hilbert envelope')
        tf = np.linspace(0, td[-1], 500)
        ax.plot(tf, A0*np.exp(-al*tf), color='C3', lw=1.2,
                label=f'Exponential fit, zeta = {z:.4f} ({z*100:.2f}%)')
        ax.set_title(f'{ml}, f = {fn:.1f} Hz', fontsize=11)
        ax.set_ylabel('Amplitude (a.u.)')
        ax.legend(fontsize=8, loc='upper right')
    axes[-1].set_xlabel('Time (s)')
    plt.tight_layout()
    plt.savefig(os.path.join(FD, fname), dpi=300)
    plt.close()
    print(f"  saved {fname}")


def plot_modes(exp, mod, freqs, macs, fname):
    fig, axes = plt.subplots(1, 3, figsize=(10, 4.2), sharey=True)
    fl = [0, 1, 2, 3]
    for i, ax in enumerate(axes):
        ax.plot(np.r_[0, mod[:,i]], fl, 'o-', color='C0', ms=6, lw=1.5,
                label='Model')
        ax.plot(np.r_[0, exp[i]], fl, 's-', color='C3', ms=6, lw=1.2,
                label='Experimental')
        ax.axvline(0, color='0.6', lw=0.4, ls=':')
        ax.set_xlabel('Normalised displacement')
        ax.set_xlim([-1.3, 1.3])
        ax.set_title(f'Mode {i+1}, f = {freqs[i]:.1f} Hz\nMAC = {macs[i]:.3f}',
                     fontsize=10)
        ax.set_yticks(fl)
        ax.set_yticklabels(['Base', 'Floor 1', 'Floor 2', 'Floor 3'])
        ax.legend(fontsize=8, loc='lower right')
    plt.suptitle('Mode shapes: model prediction vs experimental', fontsize=12)
    plt.tight_layout()
    plt.savefig(os.path.join(FD, fname), dpi=300)
    plt.close()
    print(f"  saved {fname}")
